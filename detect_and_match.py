import cv2
import time
import csv
import os
from openpyxl import Workbook, load_workbook
from datetime import datetime

face_cascade = cv2.CascadeClassifier("haarcascade_frontalface_default.xml")
recognizer = cv2.face.LBPHFaceRecognizer_create()
recognizer.read("trained_model.yml")

labels = {}
with open("labels.csv", "r") as f:
    reader = csv.reader(f)
    for row in reader:
        labels[int(row[0])] = row[1]

log_file = "face_log.xlsx"
if not os.path.exists(log_file):
    wb = Workbook()
    ws = wb.active
    ws.append(["Timestamp", "Person"])
    wb.save(log_file)

def log_face(person):
    wb = load_workbook(log_file)
    ws = wb.active
    ws.append([datetime.now().strftime("%Y-%m-%d %H:%M:%S"), person])
    wb.save(log_file)

cap = cv2.VideoCapture(0)
last_prediction_time = 0
last_label = "No Face"
last_color = (0, 0, 255)
last_logged_person = None
prediction_timeout = 2

while True:
    ret, frame = cap.read()
    if not ret:
        break

    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    faces = face_cascade.detectMultiScale(gray, 1.3, 5)
    current_time = time.time()
    face_found = False

    for (x, y, w, h) in faces:
        face_found = True
        face = gray[y:y+h, x:x+w]
        label, confidence = recognizer.predict(face)

        if confidence < 70:
            name = labels.get(label, "Unknown")
            color = (0, 255, 0)
        else:
            name = "Unknown"
            color = (0, 0, 255)

        last_label = name
        last_color = color
        last_prediction_time = current_time

        cv2.rectangle(frame, (x, y), (x+w, y+h), color, 2)
        cv2.putText(frame, name, (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.9, color, 2)

        if name != "Unknown" and name != last_logged_person:
            log_face(name)
            last_logged_person = name

    if not face_found and current_time - last_prediction_time < prediction_timeout:
        cv2.putText(frame, last_label, (30, 30), cv2.FONT_HERSHEY_SIMPLEX, 1, last_color, 2)

    cv2.imshow("Face Match", frame)
    if cv2.waitKey(1) & 0xFF == ord("q"):
        break

cap.release()
cv2.destroyAllWindows()
